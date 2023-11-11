import openpyxl
# Crear libro de Excel
lista_vehiculos = openpyxl.Workbook()

# Seleccionar la hoja
hoja = lista_vehiculos.active

# encabezado
hoja['A1'] = "codigo"
hoja['B1'] = "Marca"
hoja['C1'] = "precio"
hoja['D1'] = "Kilometraje"

# Listas

vehiculo = [
    ("001", "Mazda", "20000", 3000),
    ("002", "toyota", "40000", 3000),
    ("003", "susuky", "90000", 3000),
    ("004", "Hilux", "70000", 2000),
    ("005", "Tesla", "1000000", 100000),
    ("005", "Tesla", "1000000", 100000),
]

# Agregar datos a la hoja
for row, data in enumerate(vehiculo, start=2):
    hoja.cell(row=row, column=1, value=data[0])
    hoja.cell(row=row, column=2, value=data[1])
    hoja.cell(row=row, column=3, value=data[2])
    hoja.cell(row=row, column=4, value=data[3])

# Guardar el archivo Excel
lista_vehiculos.save("lista_vehiculos.xlsx")

# Cerrar el archivo Excel
lista_vehiculos.close()
